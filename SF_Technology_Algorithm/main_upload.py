from core.datahandle import DataHandler
from core.model import bayes_optimization, model_evaluate
from core.config import Config
import pandas as pd
from openpyxl import load_workbook
import os

upload_day = '1130'

"""
将历史数据和新数据融合
"""
data_input = DataHandler('data/quantity_train.csv', id_column = Config.id_column, time_column = Config.time_column,
                         target_column = Config.target_column)
his_handled_data = data_input.data_transform()
pre_handled_data = DataHandler.predict_handle('data/{}_predict_input.csv'.format(upload_day))

unfeature_df = pd.concat([his_handled_data, pre_handled_data], axis = 0)
unfeature_df.sort_values(by = [Config.id_column, Config.time_column], ascending = (True,True), inplace = True)
unfeature_df.reset_index(drop = True, inplace = True)

"""
预测提交结果
"""
Config.objective = 'regression'
upload_predict_results = pd.DataFrame()

for predict_target in Config.target_column:  # pai_num, shou_num
    Config.predict_target = predict_target
    with open('params/{}_params.txt'.format(predict_target), "r") as f:
        params = f.read()
    params = eval(params)
    print('XXXX:')
    print(params)
    test_mape, test_mape_result, predict_results, feature_importance_df, model, train_results \
                            = model_evaluate(unfeature_df, params = params, test_rounds = 1, test_period = 3)
    predict_results['date_str'] = '2020' + predict_results['month'].apply(lambda x: str(x)) + \
                                                                         predict_results['day'].apply(lambda x: str(x))
    dt_format = '%Y%m%d'
    predict_results['date_str'] = pd.to_datetime(predict_results['date_str'], format=dt_format)
    predict_results = predict_results.loc[:, ['date_str', 'zone_id', 'predict']]
    predict_results.columns = ['date_str', 'zone_id', predict_target]

    if upload_predict_results.empty:
        upload_predict_results = predict_results
    else:
        upload_predict_results = pd.merge(upload_predict_results, predict_results, how = 'left',
                                                                                    on = ['date_str', 'zone_id'] )
upload_predict_results.sort_values(by=['date_str','zone_id'], ascending=(True,True), inplace=True)
upload_predict_results.to_csv('upload_results/{}_forecast.csv'.format(upload_day), index = False)

"""
优化输入结果
"""
Config.objective = 'quantile'
upload_arrange_result = pd.DataFrame()
columns_name = ['时间','区域id','派件量','收件量','派件量预测','收件量预测']

for predict_target in Config.target_column:
    Config.predict_target = predict_target
    with open('params/{}_params.txt'.format(predict_target), "r") as f:
        params = f.read()
    params = eval(params)

    for i in range(75,86):
        Config.alpha = i/100
        print('YYYY:')
        print(params)
        test_mape, test_mape_result, arrange_results, feature_importance_df, model, train_results \
                                = model_evaluate(unfeature_df, params = params, test_rounds=1, test_period=3)
        arrange_results.sort_values(by=[Config.id_column, 'month','day'], ascending=(True, True, True),
                                                                                        inplace=True, ignore_index=True)
        arrange_results[columns_name[0]] = '2020' + '/' +  arrange_results['month'].apply(lambda x : str(x)) + '/' + \
                                                                        arrange_results['day'].apply(lambda x : str(x))
        arrange_results[columns_name[1]] = arrange_results[Config.id_column]

        if predict_target == 'pai_num':
            column = '派件量预测_{}'.format(str(i))
            arrange_results[column] = arrange_results['predict'].apply(lambda x : round(x))
            result_unit = arrange_results.loc[:, [columns_name[0], columns_name[1], column]]

            if upload_arrange_result.empty:
                upload_arrange_result = result_unit
            else:
                upload_arrange_result = pd.merge(upload_arrange_result, result_unit,how = 'left', on = [columns_name[0],columns_name[1]])
        else:
            column = '收件量预测_{}'.format(str(i))
            arrange_results[column] = arrange_results['predict'].apply(lambda x : round(x))
            result_unit = arrange_results.loc[:, [columns_name[0], columns_name[1], column]]

            if upload_arrange_result.empty:
                upload_arrange_result = result_unit
            else:
                upload_arrange_result = pd.merge(upload_arrange_result, result_unit, how='left', on = [columns_name[0],columns_name[1]])

# 预测信息sheet
pai_real = pd.Series([None for i in range(0,len(upload_predict_results))], name=columns_name[2])
shou_real = pd.Series([None for i in range(0,len(upload_predict_results))], name=columns_name[3])
upload_predict_results.sort_values(by=['zone_id','date_str'], ascending=(True,True), inplace=True, ignore_index=True)
upload_predict_results['date_str'] = upload_predict_results['date_str'].apply(lambda x : x.strftime('%Y/%m/%d'))
upload_predict_results['date_str'] = upload_predict_results['date_str'].apply(lambda x : x.split('/')[0]+'/'+x.split('/')[1][1]+'/'+x.split('/')[2])
upload_predict_results.columns = [columns_name[0], columns_name[1], columns_name[4], columns_name[5]]
arrange_predict = pd.concat([upload_predict_results, pai_real, shou_real], axis = 1)
print(arrange_predict)
arrange_predict = arrange_predict.loc[:,columns_name]

# 优化结果输出
path = 'upload_results/{}_data_input_level.xlsx'.format(upload_day)
wb = load_workbook(path)
ws1 = wb["预测信息"]
ws2 = wb["样本区间"]
wb.remove(ws1)
wb.remove(ws2)
wb.save(path)

writer = pd.ExcelWriter(path, engine='openpyxl')
book = load_workbook(path)
writer.book = book
arrange_predict.to_excel(writer, sheet_name='预测信息', index = False)
upload_arrange_result.to_excel(writer, sheet_name='样本区间', index = False)
writer.save()





